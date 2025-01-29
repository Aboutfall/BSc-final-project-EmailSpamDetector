
from openpyxl import workbook, load_workbook
from ast import Delete, Or, Return
from operator import contains, countOf
import pandas as pd
import numpy as np
import matplotlib.pyplot as mp
import seaborn as sb
import scipy as sy
import statsmodels as sm

import sklearn as sk
from sklearn.feature_selection import SelectKBest, chi2
from sklearn import model_selection, metrics
from sklearn.metrics import recall_score, make_scorer, precision_score,f1_score,precision_score
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.pipeline import Pipeline
from sklearn.naive_bayes import MultinomialNB, GaussianNB
from sklearn.ensemble import RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier




class FinalProject(object):
    #FUNCTIONS
    #========================================
    #Tokenization of letters
    def tokenization (em1):
    #initializing variables
      emnum = 0
      uniqnum = 0
      library=[]
      matrix = []
      ste = 0
      wletters = ["\x01","\x16","\x1b","\x14","\x12","\x02","\x03","\x04","\x05","\x06","\x07","\x08","\x09","\x0a","\x0b","\x0c","\x0d","\x0e","\x0f","\x11","\x13"]

     #Goes through all the emails
      for i in em1:
          #Keeps track of words in emails
          wordnum = 0
          #Visualising how many have been done
          print(emnum+1)
          #Used to ensure that the word is unique in the email
          emstart = len(matrix)
          # j = all the words within the email
          for j in i:
              #To check if its unique
              unique = True
              # k = 0 to library length which is a library of unique words
              for k in range(len(library)):
                 #used to check if the word is unique in library
                 if j == library[k]:
                     
                     unique = False
                     ste = k
                     # matrix values which [which email | the position within library | amount]
                     m = [emnum,ste,1]
                     break
              # check to see if its unique
              if unique: 
                  # matrix values for [email number | unique number | amount which will always be 1]
                  m = [emnum,uniqnum,1]
                  matrix.append(m)
                  #removes all hexadecimal values which show as string 
                  if wletters.__contains__(j) == False :
                    library.append(j)
                    uniqnum = uniqnum+1
                 
              else:
                  # Used to go through current matrix of that email to decide whether its unique within email or not 
                  ms = len(matrix)
                  if emstart-ms == 0:
                      unique = True
                  # gets values from email start to end of matrix 
                  for l in range(emstart,ms):
                    # checking if unique number is equal to the number in the library which is: ste 
                    if matrix[l][1] == ste:
                        #Adds 1 to the amount 
                     matrix[l][2] = matrix[l][2]+1
                     unique = False
                     break
                 # Otherwise its unique
                    else:
                     unique =True

                  if unique:
                      # this can be the m from where the unique value already exists or its completely a unique word
                      matrix.append(m)
              #
              wordnum = wordnum+1
          emnum = emnum+1

      return matrix,library 
    # write to XLSX  
    def writeToCSV(lib,mat,sOrH):
        #Uses workbook to write into individual cells
        wk = load_workbook(r"C:\Users\Christian\Final project\emailExtract.xlsx")
        ws = wk.active
        count = 1
        
        ws.title = "Email data"
        
        print("Writing to XLSX")
        print("-----------------")
        
        
        #first writes all the entries of the library in terms of 'frequency_of' as column headers
        for i in lib:
                word = "frequency_of_ " + i
                
                ws.cell(row = 1,column = count,value = word)
                count = count +1
        # writes the frequency of the word as data into cells with the column being the unique word and the row being the email
        for i in mat:
                    
                try:
                    ws.cell(row = i[0]+2,column = i[1]+1,value = i[2])
                except :
                       pass
        count = 2
        #Writes the classification of the email in either 1 for Spam or 0 for ham 
        for i in sOrH[0:5000]:
            ws.cell(row = 1,column =len(lib)+1,value = "Spam_Ham") 
            if i == "spam":
                ws.cell(row = count,column = len(lib)+1,value = 1)
            else:
                ws.cell(row = count,column = len(lib)+1,value = 0)
            count = count +1
        
        wk.save(r"C:\Users\Christian\Final project\emailExtract.xlsx")
       
    # Word removal makes a new array which seperates letters with commas
    def wordRemoval (arr1):
      
        ema = []
        #split lines removes all the values in string such as /n or r/n which exist
        ema = arr1.splitlines()
        words = []
        # list of all the numbers to be identified 
        numbers = ["1","2","3","4","5","6","7","8","9","0"]
        # j being each email in array
        for j in ema:        
            #count and count2 being used to store the location of: the start of the previous word 
            #and 2 where the next one ends respectively
            count = 0
            count2 = 0
            i = 0
            # i being the letters/numbers in j
            for i in (j):
                #removes all the numbers in the email
               if numbers.__contains__(i):
                count = count2+ 1 
               else:
                   #Finds when theres a space in the text signifying a word
                if i == " " :
                    #ensures that it isn't nothing between the counts
                   if j[count:count2] != " ":
                    
                     words.append(j[count:count2])
                     count = count2 +1
               count2 = count2 +1
        return(words)
  
    #Removal of stopwords
    def removeStopwords (ema):
        # list of some of the stop words in the english dictionary
        listOfStopWords = [""," ","y","q","w","e","r","t","u","i","o","p","s","d","f","g","h","j","k","l","z","x","c","v","b","n","m","a","ect","ve","re","nt","about","after","against","am","and","are","as","above","again","all","an","at","aren't","any","be","been","being","between","but","because","before","both","below","by","couldnt","can","cant","could","cannot","didnt","dont","during","does","do","doesnt","down","doing","did","each","few","further","for","from","hi","had","hadnt","has","have","he","hed","hes","her","heres","hers","his","hows","here","how","hasnt","havent","him","herself","himself","hell","having","http","https","ill","i","id","im","ive","if","in","is","isnt","its","it","itself","into","lets","me","my","most","more","mustnt","myself","not","no","nor","on","once","only","our","or","ours","other","out","ought","over","of","off","own","ourselves","she","shes","shell","same","shant","so","shed","such","some","shouldnt","should","them","theyll","they","this","themselves","than","then","through","theyve","too","theyd","there","those","thats","theres","to","the","theirs","their","that","these","theyre","until","under","up","very","was","were","wasnt","wed","were","well","werent","weve","what","www","we","whats","when","whens","where","wheres","which","while","who","whos","whom","why","whys","with","wont","would","wouldnt","you","youd","youll","youre","youve","your","yours","yourself","yourselves"]
        #output list
        em1 = []
        count = 0 
        #goes through the email and removes all of the stop words using the count method
        for i in ema:
            #finds if the word is in the list 
            if listOfStopWords.__contains__(i):
               count = count+1
            else:
                em1.append(i)
        return(em1,count)

    #Removal of specialChar
    def removeSpecChar (em1):
        newarr = ""
        # Used to keep track where in the email it is so it can remove symbols
        cnt = 0 
        # sets symbols to be removed 
        setOfSpecialChar = set(r"""``{[]}!@\:;'"<>\()?_,-+%^&*=~#$|./""")
        
        # goes through all the letters in the email 
        for j in range(len(em1)):
            
            #This looks if the character is the same as any of the letters in the set
            if setOfSpecialChar.__contains__(em1[j]):
                #This makes a new email containing only the characters allowed
                 newarr = newarr + em1[cnt:j]
                 cnt = j+1
            else:
                #Looks to see if its empty to add the word without character back in the email 
                if em1[j] == " ":
                    newarr = newarr + em1[cnt:j+1]
                    cnt = j+1
        
        return(newarr)

    #Removal of redundant words
    def redundant_words(lib,mat):
        uniwordcount = []
        newlib = []
        newmat = []
        count = 0
        val = True
        # Goes through the matrix entries
        for i in mat:
            # Goes through uniwordcount which counts the frequency of each word in the library finding if theres a new word to be added
            for j in uniwordcount:
             if j[0] != i[1]:
               
                val = True
             else:
                val = False
                j[1] = j[1] + 1
                break
            # Adds unique word to be counted
            if val :
             uniwordcount.append([i[1],1])
        #goes through uniwordcount to find if the unique word in the entire dataset is > x
        for i in uniwordcount:
            if i[1] >  2  :
                newlib.append(lib[i[0]])
              
                for j in mat:
                    if j[1] == i[0]:
                        newmat.append([j[0],count,j[2]])
                count = count +1
        return newlib,newmat
    
    #INITIATION
    #=====================================
    # reads the csv database of email spam 
    data = pd.read_csv(r'C:\Users\Christian\Final project\spam_ham_dataset.csv',encoding = 'utf8');
  
    # Sets an array for the emails 
    arr = np.array(data)
    
    # Grabs only the emails from the csv files 
    emails = arr[:,2]
    # Array with whether email is spam or not 
    sOrH = arr[:,1]
    count = 0 
    murph = 0
    df = []

    #Used to test part of the code
    Emailextract = False
    featureSelect = False
    Classifier = True

    # EXECUTING CODE
    #==============================================
    if Emailextract :
     for i in emails:
        count = count +1 
        em1 = i
        em1 = removeSpecChar(em1)
        em1 = wordRemoval(em1)
        (em1,num) = removeStopwords(em1)
        df.append(em1)
        murph = murph +1
        print(murph)
        if count == 100:
           break
    
     print("tokenization")
     matrix,library = tokenization(df)

     print("Redundant word removal")
     library,matrix = redundant_words(library,matrix)

     writeToCSV(library,matrix,sOrH)
    
     #reads data from excel
    df = pd.read_excel(r"C:\Users\Christian\Final project\emailExtract.xlsx")
    #fills any blank space with a 0
    df = df.fillna(0)
    #seperate the data 
    features = df.drop('Spam_Ham', axis = 1)
    spam_ham = df['Spam_Ham']
    
    if featureSelect:
   
     #Chi-Square feature selection k being the number of features
     featureSelection = SelectKBest(chi2, k = 1000)

     featureSelection.fit(features,spam_ham)
     res = features.columns[featureSelection.get_support()]
   
   
     bol = True 
     count = 0
     ran = len(features.columns)
     #Goes through all the emails column names
     for j in features.columns:
      print(str(count) + " out of " + str(ran))
      #goes through all the selected features
      for i in res:
        if j != i:
            bol = True
        else:
            bol = False
            break
      count = count+1
      #drops the column if it isnt a selected word
      if bol:
       features = features.drop(j,axis = 1)
       #writes to CSV 
     features.to_csv(r'C:\Users\Christian\Final project\keywords_1000.csv', encoding = 'utf8')
     print("features collected")

    if Classifier:
     df = pd.read_csv(r'C:\Users\Christian\Final project\keywords_500.csv', encoding = 'utf8')
        
    #machine learnign algorithms
     x = df.to_numpy()
     y = spam_ham.to_numpy()
     
     #Splitting the data into train-test,    
     x_tr,x_t,y_tr,y_t = model_selection.train_test_split(x,y,test_size = 0.66, shuffle = True)
     
     rndseed = 42
    #classifiers

    # Multinomial NB
     alg1 = {}
     #Parametres
     alg1['classifier__alpha'] = [1,10,100,250]
     alg1['classifier'] = [MultinomialNB()]
   
    # RandomForest
     alg2 = {}
     #Parametres
     alg2['classifier__n_estimators'] = [10,25,50,100,250]
     alg2['classifier__max_depth'] = [5,10,20,30,None]
     alg2['classifier__class_weight'] = [None,{0:1,1:5},{0:1,1:10},{0:1,1:25}]
     alg2['classifier'] = [RandomForestClassifier(random_state=rndseed)]

     #Gaussian NB
     alg3 = {}  
     alg3['classifier'] = [GaussianNB()]

        


     #setting up pipeline for classifiers
     pipeline = Pipeline([('classifier',GaussianNB())])
     #using gridsearchCV with pipeline to find and compare best algorithm using x_train,y_train
     #making scorer specificity for Spam classification
     scorer = make_scorer(f1_score,greater_is_better=True,pos_label = 0) 
     
     trainGS = GridSearchCV(pipeline,[alg3],cv = 10,scoring = scorer,verbose = 1).fit(x_tr,y_tr)
     #using the best algorithm to predict what the test data could be.
     prediction = trainGS.predict(x_t)   
     m = metrics.confusion_matrix(prediction, y_t)
  
     #output of confusion matrix and percentage. 
     
     print(trainGS.best_score_)
     print(trainGS.best_params_)  
     print(m)
     print("-----------------")
     print(metrics.classification_report(prediction,y_t))
     
pass 